import streamlit as st
import openpyxl
from openpyxl.styles import PatternFill, Font, Color
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
import pandas as pd
from io import BytesIO

st.set_page_config(page_title="CF Rules Applicator", page_icon="📊", layout="wide")

def parse_rgb(rgb_string):
    """Parse RGB string like '255,0,0' to hex"""
    if not rgb_string or rgb_string.strip() == '':
        return None
    cleaned = rgb_string.replace("'", "").strip()
    parts = cleaned.split(',')
    if len(parts) == 3:
        try:
            r, g, b = [int(p.strip()) for p in parts]
            return f"{r:02X}{g:02X}{b:02X}"
        except:
            return None
    return None

def color_name_to_hex(color_name):
    """Convert color name to hex"""
    color_map = {
        'Aqua': '00FFFF', 'Black': '000000', 'Blue': '0000FF', 'BlueGray': '666699',
        'BrightGreen': '00FF00', 'Brown': '993300', 'DarkBlue': '000080', 'DarkGreen': '003300',
        'DarkRed': '800000', 'DarkYellow': '808000', 'DarkTeal': '003366', 'Gold': 'FFC000',
        'Green': '008000', 'Gray25': 'C0C0C0', 'Gray40': '969696', 'Gray50': '808080',
        'Gray80': '333333', 'Grey': '808080', 'Gray': '808080', 'Indigo': '333399',
        'Lavender': 'CC99FF', 'LightBlue': '3366FF', 'LightGray': 'C0C0C0', 'LightGreen': 'CCFFCC',
        'LightOrange': 'FF9900', 'LightTurquoise': 'CCFFFF', 'LightYellow': 'FFFF99',
        'Lime': '99CC00', 'OliveGreen': '333300', 'Orange': 'FF6600', 'PaleBlue': '99CCFF',
        'Pink': 'FF00FF', 'Plum': '993366', 'Red': 'FF0000', 'Rose': 'FF99CC',
        'SeaGreen': '339966', 'SkyBlue': '00CCFF', 'Tan': 'FFCC99', 'Teal': '008080',
        'Turquoise': '00FFFF', 'Violet': '800080', 'White': 'FFFFFF', 'Yellow': 'FFFF00'
    }
    return color_map.get(color_name)

def get_color_hex(color_name, rgb_string):
    """Get hex color from name or RGB string"""
    if rgb_string and rgb_string.strip():
        hex_color = parse_rgb(rgb_string)
        if hex_color:
            return hex_color
    
    if color_name and color_name.strip() and color_name not in ['RGB', 'NoColor', '']:
        hex_color = color_name_to_hex(color_name)
        if hex_color:
            return hex_color
    
    return None

def column_letter_to_index(col):
    """Convert column letter to index (A=1, B=2, etc.)"""
    index = 0
    for char in col:
        index = index * 26 + (ord(char) - ord('A') + 1)
    return index

def apply_cf_rules(rules_file, target_file, row_number_input, first_row, last_row, mode):
    """Apply conditional formatting rules to Excel file"""
    
    # Determine which row number to use
    if row_number_input and row_number_input.strip():
        row_number = row_number_input.strip()
    elif first_row:
        row_number = str(first_row)
    else:
        st.error("❌ Please enter either a Row Number or First Row!")
        return None
    
    # Load the rules file
    wb_rules = openpyxl.load_workbook(rules_file)
    
    if 'CF Rules' not in wb_rules.sheetnames:
        st.error("❌ 'CF Rules' sheet not found in the uploaded file!")
        return None
    
    rules_sheet = wb_rules['CF Rules']
    
    # Read all rules
    rules = []
    for row in rules_sheet.iter_rows(min_row=2, values_only=True):
        if row[0]:  # If Start Column exists
            rules.append({
                'start_col': row[0],
                'end_col': row[1],
                'formula': row[2] if row[2] else '',
                'stop_if_true': row[3] == 'Y' if row[3] else False,
                'bg_color': row[4] if len(row) > 4 else '',
                'bg_rgb': row[5] if len(row) > 5 else '',
                'font_color': row[6] if len(row) > 6 else '',
                'font_rgb': row[7] if len(row) > 7 else '',
                'number_format': row[8] if len(row) > 8 else '',
                'worksheet': row[9] if len(row) > 9 else 'Sheet1'
            })
    
    if not rules:
        st.error("❌ No rules found in the CF Rules sheet!")
        return None
    
    st.success(f"✅ Loaded {len(rules)} rules from the file")
    
    # Create or load workbook based on mode
    if mode == "Update Existing File":
        if not target_file:
            st.error("❌ Please upload a target file to update!")
            return None
        wb_new = openpyxl.load_workbook(target_file)
        st.info("📝 Updating existing file...")
    else:  # Create New File
        wb_new = openpyxl.Workbook()
        wb_new.remove(wb_new.active)  # Remove default sheet
        st.info("📄 Creating new file...")
    
    # Group rules by worksheet
    rules_by_sheet = {}
    for rule in rules:
        sheet_name = rule['worksheet'] if rule['worksheet'] else 'Sheet1'
        if sheet_name not in rules_by_sheet:
            rules_by_sheet[sheet_name] = []
        rules_by_sheet[sheet_name].append(rule)
    
    # Determine row range for application
    if first_row and last_row:
        start_row = first_row
        end_row = last_row
    else:
        start_row = 1
        end_row = 1000
    
    # Create sheets and apply rules
    progress_bar = st.progress(0)
    status_text = st.empty()
    
    sheet_count = 0
    total_sheets = len(rules_by_sheet)
    rules_applied = 0
    rules_skipped = 0
    
    for sheet_name, sheet_rules in rules_by_sheet.items():
        status_text.text(f"Processing sheet: {sheet_name}")
        
        # Get or create worksheet
        if sheet_name in wb_new.sheetnames:
            ws = wb_new[sheet_name]
            # Clear existing conditional formatting
            ws.conditional_formatting._cf_rules.clear()
        else:
            ws = wb_new.create_sheet(sheet_name)
        
        # Apply each rule
        for rule in sheet_rules:
            try:
                # Replace @ROW@ in formula
                formula = rule['formula'].replace('@ROW@', row_number) if rule['formula'] else ''
                
                if not formula:
                    rules_skipped += 1
                    continue
                
                # Get colors
                bg_hex = get_color_hex(rule['bg_color'], rule['bg_rgb'])
                font_hex = get_color_hex(rule['font_color'], rule['font_rgb'])
                
                # Only create rule if we have at least one color
                if not bg_hex and not font_hex:
                    rules_skipped += 1
                    continue
                
                # Create DifferentialStyle - only set background color, leave pattern default
                fill = PatternFill(bgColor=Color(rgb=bg_hex)) if bg_hex else None
                font = Font(color=Color(rgb=font_hex)) if font_hex else None
                
                dxf = DifferentialStyle(fill=fill, font=font)
                
                # Create the range string
                range_string = f"{rule['start_col']}{start_row}:{rule['end_col']}{end_row}"
                
                # Create Rule with DifferentialStyle
                cf_rule = Rule(
                    type='expression',
                    dxf=dxf,
                    formula=[formula],
                    stopIfTrue=rule['stop_if_true']
                )
                
                # Add rule to worksheet
                ws.conditional_formatting.add(range_string, cf_rule)
                rules_applied += 1
                
            except Exception as e:
                st.warning(f"⚠️ Error applying rule for {rule['start_col']}:{rule['end_col']} - {str(e)}")
                rules_skipped += 1
                continue
        
        sheet_count += 1
        progress_bar.progress(sheet_count / total_sheets)
    
    status_text.text(f"✅ Processed {rules_applied} rules ({rules_skipped} skipped)")
    progress_bar.progress(1.0)
    
    # Save to BytesIO
    output = BytesIO()
    wb_new.save(output)
    output.seek(0)
    
    return output, rules_by_sheet, rules, rules_applied, rules_skipped, row_number, start_row, end_row

# Streamlit UI
st.title("📊 Conditional Formatting Rules Applicator")
st.markdown("Upload your CF Rules Excel file and apply conditional formatting to a workbook")

st.divider()

# Mode selection
mode = st.radio(
    "📋 Select Mode:",
    ["Create New File", "Update Existing File"],
    horizontal=True,
    help="Choose whether to create a new file or update an existing one"
)

st.divider()

# File uploaders
col1, col2 = st.columns(2)

with col1:
    rules_file = st.file_uploader(
        "📁 1. Upload CF Rules File (required)",
        type=['xlsx'],
        help="Upload the Excel file with 'CF Rules' sheet",
        key="rules_file"
    )

with col2:
    if mode == "Update Existing File":
        target_file = st.file_uploader(
            "📁 2. Upload Target File (required)",
            type=['xlsx'],
            help="Upload the Excel file you want to apply CF rules to",
            key="target_file"
        )
    else:
        target_file = None
        st.info("🆕 A new file will be created with empty sheets")

if rules_file:
    st.divider()
    st.subheader("⚙️ Configuration")
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        first_row = st.number_input(
            "🔢 First Row",
            min_value=1,
            value=None,
            step=1,
            help="Starting row number for the range (e.g., 13 for row 13)",
            placeholder="e.g., 13"
        )
    
    with col2:
        last_row = st.number_input(
            "🔢 Last Row",
            min_value=1,
            value=None,
            step=1,
            help="Ending row number for the range (e.g., 15 for row 15)",
            placeholder="e.g., 15"
        )
    
    with col3:
        row_number_input = st.text_input(
            "🔄 Row Number for @ROW@",
            value="",
            help="Enter row number to replace @ROW@ in formulas. If empty, uses First Row value.",
            placeholder="Leave empty to use First Row"
        )
    
    # Show helpful message
    if not row_number_input and first_row:
        st.info(f"ℹ️ @ROW@ will be replaced with: **{first_row}**")
    elif row_number_input:
        st.info(f"ℹ️ @ROW@ will be replaced with: **{row_number_input}**")
    
    if first_row and last_row:
        st.success(f"✅ Rules will be applied to range: **Column:Row {first_row} to Column:Row {last_row}**")
        st.caption(f"Example: If a rule has columns AE:AJ, it will apply to **AE{first_row}:AJ{last_row}**")
    
    st.divider()
    
    # Check if we can proceed
    can_proceed = True
    if mode == "Update Existing File" and not target_file:
        st.warning("⚠️ Please upload a target file to update")
        can_proceed = False
    
    if can_proceed and st.button("🚀 Generate Excel File with CF Rules", type="primary", use_container_width=True):
        with st.spinner("Processing rules and generating file..."):
            result = apply_cf_rules(rules_file, target_file, row_number_input, first_row, last_row, mode)
            
            if result:
                output_file, rules_by_sheet, rules, rules_applied, rules_skipped, row_used, start_row, end_row = result
                
                # Download button
                filename_suffix = "Updated" if mode == "Update Existing File" else "New"
                st.download_button(
                    label="⬇️ Download Generated Excel File",
                    data=output_file,
                    file_name=f"CF_{filename_suffix}_Row{row_used}_Range{start_row}-{end_row}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary",
                    use_container_width=True
                )
                
                st.divider()
                
                # Show summary
                st.subheader("📋 Summary")
                col1, col2, col3, col4 = st.columns(4)
                
                with col1:
                    st.metric("Mode", mode)
                with col2:
                    st.metric("Rules Applied", rules_applied)
                with col3:
                    st.metric("Rules Skipped", rules_skipped)
                with col4:
                    st.metric("Worksheets", len(rules_by_sheet))
                
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric("@ROW@ Value", row_used)
                with col2:
                    st.metric("First Row", start_row)
                with col3:
                    st.metric("Last Row", end_row)
                
                if rules_skipped > 0:
                    st.info(f"ℹ️ {rules_skipped} rules were skipped (no formula or no colors specified)")
                
                # Show rules by sheet
                st.subheader("📊 Rules Applied by Worksheet")
                
                for sheet_name, sheet_rules in rules_by_sheet.items():
                    with st.expander(f"📄 {sheet_name} ({len(sheet_rules)} rules)"):
                        df_data = []
                        for rule in sheet_rules:
                            formula = rule['formula'].replace('@ROW@', row_used) if rule['formula'] else '-'
                            bg_display = rule['bg_rgb'] if rule['bg_rgb'] else rule['bg_color'] if rule['bg_color'] else '-'
                            font_display = rule['font_rgb'] if rule['font_rgb'] else rule['font_color'] if rule['font_color'] else '-'
                            range_display = f"{rule['start_col']}{start_row}:{rule['end_col']}{end_row}"
                            
                            df_data.append({
                                'Range': range_display,
                                'Formula': formula[:60] + '...' if len(formula) > 60 else formula,
                                'BG Color': bg_display,
                                'Font Color': font_display,
                                'Stop If True': '✓' if rule['stop_if_true'] else ''
                            })
                        
                        df = pd.DataFrame(df_data)
                        st.dataframe(df, use_container_width=True, hide_index=True)

else:
    st.info("👆 Please upload a CF Rules Excel file to get started")
st.divider()
st.caption("💡 Tip: Use 'Update Existing File' mode to add CF rules to files with existing data!")